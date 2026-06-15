"""
ADA Application Portal — Enhanced Flask Backend
- Unique Application ID (ADA202600001)
- Auto submission date/time
- Multi-sheet Excel (Applicants, Family Details, Education Details)
- Consolidated PDF (all details + undertaking appended at end)
- Photo compression (PNG/JPG/JPEG → max 20KB)
- PDF compression (max 100KB)
- Applicant can download PDF after submission
- Admin: search by App ID, view date/time
"""

import os, uuid, hashlib, io
from datetime import datetime
from flask import (Flask, render_template, request, redirect,
                   url_for, flash, send_from_directory, jsonify, session, send_file)
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, HRFlowable, Image as RLImage, PageBreak)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT
from reportlab.pdfgen import canvas as rl_canvas
from PIL import Image as PILImage
import pikepdf

# ── Config ─────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key  = os.environ.get("SECRET_KEY", os.urandom(32))
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SECURE"]   = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["MAX_CONTENT_LENGTH"]      = 5 * 1024 * 1024

ADMIN_EMAIL     = os.environ.get("ADMIN_EMAIL", "ada.admin@portal.in")
ADMIN_PASSWORD  = os.environ.get("ADMIN_PASSWORD", "ADA@Admin2026")
ADMIN_PASS_HASH = hashlib.sha256(ADMIN_PASSWORD.encode()).hexdigest()


#Find a folder where app.py is located
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))

UPLOAD_PDF_DIR  = os.path.join(BASE_DIR, "uploads", "pdfs")
UPLOAD_PHO_DIR  = os.path.join(BASE_DIR, "uploads", "photos")
UPLOAD_GEN_DIR  = os.path.join(BASE_DIR, "uploads", "generated")
SAMPLE_DIR      = os.path.join(BASE_DIR, "samples")
EXCEL_PATH      = os.path.join(BASE_DIR, "APPLICATION.xlsx")

@app.route("/where")
def where():
    return f"""
    BASE_DIR: {BASE_DIR}<br>
    Photos: {UPLOAD_PHO_DIR}<br>
    PDFs: {UPLOAD_PDF_DIR}<br>
    Generated PDFs: {UPLOAD_GEN_DIR}<br>
    Excel: {EXCEL_PATH}
    """

ALLOWED_PDF = {"pdf"}
ALLOWED_IMG = {"jpg", "jpeg", "png"}

for d in [UPLOAD_PDF_DIR, UPLOAD_PHO_DIR, UPLOAD_GEN_DIR, SAMPLE_DIR]:
    os.makedirs(d, exist_ok=True)

# ── Application ID ──────────────────────────────────────────
def generate_app_id():
    year = datetime.now().year
    if not os.path.exists(EXCEL_PATH):
        return f"ADA{year}00001"
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = "Application" if "Application" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        max_num = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[57] if len(row) > 57 else None
            if val and str(val).startswith(f"ADA{year}"):
                try:
                    num = int(str(val)[7:])
                    max_num = max(max_num, num)
                except: pass
        return f"ADA{year}{max_num+1:05d}"
    except:
        return f"ADA{year}00001"

# ── File helpers ────────────────────────────────────────────
def allowed_file(filename, allowed_set):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed_set


def unique_filename(ext):
    return f"{uuid.uuid4().hex}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"

# ── Excel Setup ─────────────────────────────────────────────
def ensure_excel():
    """Ensure APPLICATION.xlsx exists with all 4 sheets. Seed from base template if missing."""
    import shutil as _sh
    base = os.path.join(BASE_DIR, "APPLICATION_BASE.xlsx")
    if not os.path.exists(EXCEL_PATH):
        if os.path.exists(base):
            _sh.copy(base, EXCEL_PATH)
            _ensure_extra_sheets()
        else:
            _build_excel_fresh()
        return
    # File exists — make sure all 4 sheets present
    wb = openpyxl.load_workbook(EXCEL_PATH)
    needed = ["Application","Admin","Family Details","Education Details"]
    missing = [s for s in needed if s not in wb.sheetnames]
    if missing:
        for s in missing:
            wb.create_sheet(s)
        wb.save(EXCEL_PATH)
        _ensure_extra_sheets()

def _ensure_extra_sheets():
    """Add headers to Family Details and Education Details sheets if blank."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    _add_family_headers(wb)
    _add_education_headers(wb)
    wb.save(EXCEL_PATH)

def _hdr(ws, vals, row=1, fill_color="1565C0"):
    hf = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    hfl = PatternFill("solid", fgColor=fill_color)
    thin = openpyxl.styles.borders.Side(style="thin", color="B0BEC5")
    bdr  = openpyxl.styles.borders.Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, v in enumerate(vals, 1):
        c = ws.cell(row, i, v or "")
        if v:
            c.font=hf; c.fill=hfl; c.alignment=ctr; c.border=bdr
    ws.row_dimensions[row].height = 28

def _add_family_headers(wb):
    ws = wb["Family Details"]
    if ws.cell(1,1).value:
        return
    headers = [
        "APPLICATION ID","APPLICANT NAME",
        "FATHER NAME","FATHER OCCUPATION","FATHER MOBILE",
        "MOTHER NAME","MOTHER OCCUPATION","MOTHER MOBILE",
        "SIBLING 1 NAME","SIBLING 1 RELATION","SIBLING 1 OCCUPATION","SIBLING 1 MOBILE",
        "SIBLING 2 NAME","SIBLING 2 RELATION","SIBLING 2 OCCUPATION","SIBLING 2 MOBILE",
        "SIBLING 3 NAME","SIBLING 3 RELATION","SIBLING 3 OCCUPATION","SIBLING 3 MOBILE",
        "SIBLING 4 NAME","SIBLING 4 RELATION","SIBLING 4 OCCUPATION","SIBLING 4 MOBILE",
        "NUMBER OF SIBLINGS","SIBLING DETAILS"
    ]
    _hdr(ws, headers, 1)
    widths = [16,22,22,18,14,22,18,14,20,16,18,14,20,16,18,14,20,16,18,14,20,16,18,14,12,30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width = w
    ws.freeze_panes = "C2"

def _add_education_headers(wb):
    ws = wb["Education Details"]
    if ws.cell(1,1).value:
        return
    # Row 1: merged group headers
    groups = [("APPLICATION ID",1),("APPLICANT NAME",1),("BRANCH",1),
              ("10th STANDARD",4),("INTERMEDIATE / DIPLOMA",4),("DEGREE / B.TECH",4),
              ("MASTERS / M.TECH / MBA",4),("PhD / ADDITIONAL",4)]
    col = 1
    hf = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    hfl = PatternFill("solid", fgColor="1565C0")
    thin = openpyxl.styles.borders.Side(style="thin", color="B0BEC5")
    bdr  = openpyxl.styles.borders.Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for title, span in groups:
        c = ws.cell(1, col, title)
        c.font=hf; c.fill=hfl; c.alignment=ctr; c.border=bdr
        if span > 1:
            ws.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+span-1)
        col += span
    ws.row_dimensions[1].height = 28
    # Row 2: sub-headers
    sf = Font(bold=True, color="0D47A1", name="Arial", size=9)
    sfl = PatternFill("solid", fgColor="BBDEFB")
    sub = ["APP ID","NAME","BRANCH"]
    for _ in range(5):
        sub += ["SCHOOL/INSTITUTION","BOARD/UNIVERSITY","YEAR OF PASSING","PERCENTAGE/CGPA"]
    for i, v in enumerate(sub, 1):
        c = ws.cell(2, i, v)
        c.font=sf; c.fill=sfl; c.alignment=ctr; c.border=bdr
    ws.row_dimensions[2].height = 22
    for i, w in enumerate([14,22,18]+[26,22,14,15]*5, 1):
        ws.column_dimensions[ws.cell(2,i).column_letter].width = w
    ws.freeze_panes = "D3"

def _build_excel_fresh():
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Application"

    row1_vals = [
        "S No","Name","Qualification B.E/B.Tech/M.E/M.Tech","Branch","DOB","College Name",
        "Semester","CGPA","DESIGNATION","AADHAAR NUMBER","GENDER",
        "MARITAL STATUS","HEALTH ISSUES","PERMANENT ADDRESS","PRESENT ADDRESS","PREVIOUS RESIDENTIAL ADDRESS",
        "CONTACT NO. MOBILE NO","IMEI NO","LANDLINE NO","WHATSAPP NO.","FACEBOOK ID","E-MAIL ID",
        "LINKEDIN ID","INSTAGRAM ID","OTHER ID. (IF ANY)","BANK NAME","BANK ACCOUNT NO.","IFSC CODE",
        "PAN NO.","EPF NO / PF NO","ESIC NO","PROJECT TITLE","GUIDE / DIRECTORATE","AREA OF WORK",
        "DURATION FROM","DURATION TO","UNIVERSITY STUDENT REGISTRATION NO",
        "FAMILY NAME","FAMILY CONTACT","FAMILY RELATION","FAMILY AGE","FAMILY MARITAL","FAMILY OCCUPATION",
        "EXAM NAME","BOARD/UNIVERSITY","YEAR PASSING","SUBJECT","PERCENTAGE",
        # Previous Employment (4 cols)
        "PREV EMP - COMPANY NAME","PREV EMP - FROM","PREV EMP - TO","PREV EMP - CATEGORIES OF WORK",
        # Foreign Employment (5 cols)
        "FOREIGN - NAME","FOREIGN - RELATION","FOREIGN - JOB","FOREIGN - COUNTRY","FOREIGN - DURATION",
        "",   # 57 blank
        "APPLICATION ID","SUBMISSION DATE","SUBMISSION TIME",
        "PHOTO FILE","REC LETTER FILE","GENERATED PDF",
    ]
    _hdr(ws1, row1_vals, 1)
    ws1.freeze_panes = "B2"
    wb.create_sheet("Admin")
    ws2 = wb["Admin"]
    _hdr(ws2, ["S No","Name","Qualification","Branch","DOB","College Name","Semester","CGPA","From","To","APPLICATION ID"], 1)
    ws2.freeze_panes = "B2"
    wb.create_sheet("Family Details")
    wb.create_sheet("Education Details")
    wb.save(EXCEL_PATH)
    _ensure_extra_sheets()

def save_to_excel(data, app_id, sub_date, sub_time, photo_name, pdf_name, gen_pdf_name):
    ensure_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)

    thin = openpyxl.styles.borders.Side(style="thin", color="D0D0D0")
    bdr  = openpyxl.styles.borders.Border(left=thin,right=thin,top=thin,bottom=thin)
    lft  = Alignment(horizontal="left", vertical="center", wrap_text=True)
    alt1 = PatternFill("solid", fgColor="EBF5FB")
    alt2 = PatternFill("solid", fgColor="FDFEFE")

    def wr(ws, vals, fill=None):
        rn = ws.max_row + 1
        for i, v in enumerate(vals, 1):
            c = ws.cell(rn, i, v if v is not None else "")
            c.alignment = lft; c.border = bdr
            if fill: c.fill = fill
        ws.row_dimensions[rn].height = 18
      
    ws_app = wb["Application"]
    sno = max(0, ws_app.max_row - 1) + 1  # 1 header row
    fill = alt1 if sno % 2 == 1 else alt2
    app_row = [
        sno,                                    #  1  S No
        data.get("name"),                       #  2  Name
        data.get("qualification"),              #  3  Qualification
        data.get("btech_branch"),               #  4  Branch
        data.get("dob"),                        #  5  DOB
        data.get("college_name"),               #  6  College
        data.get("btech_year"),                 #  7  Semester/Year
        data.get("btech_cgpa"),                 #  8  CGPA
        data.get("designation"),                #  9  Designation
        data.get("aadhaar"),                    # 10  Aadhaar
        data.get("gender"),                     # 11  Gender
        data.get("marital_status"),             # 12  Marital Status
        data.get("health_issues"),
        data.get("permanent_address"),          # 13  Permanent Address
        data.get("present_address"),            # 14  Present Address
        data.get("previous_address"),           # 15  Previous Address
        data.get("mobile"),                     # 16  Mobile
        data.get("imei"),                       # 17  IMEI
        data.get("landline"),                   # 18  Landline
        data.get("whatsapp"),                   # 19  WhatsApp
        data.get("facebook"),                   # 20  Facebook
        data.get("email"),                      # 21  Email
        data.get("linkedin"),                   # 22  LinkedIn
        data.get("instagram"),                  # 23  Instagram
        data.get("other_id"),                   # 24  Other ID
        data.get("bank_name"),                  # 25  Bank Name
        data.get("bank_account"),               # 26  Bank Account
        data.get("ifsc"),                       # 27  IFSC
        data.get("pan"),                        # 28  PAN
        data.get("epf"),                        # 29  EPF
        data.get("esic"),                       # 30  ESIC
        data.get("project_title"),              # 31  Project Title
        data.get("guide"),                      # 32  Guide
        data.get("area_of_work"),               # 33  Area of Work
        data.get("duration_from"),              # 34  Duration From
        data.get("duration_to"),                # 35  Duration To
        data.get("university_reg"),             # 36  University Reg No
        data.get("father_name"),                # 37  Family Name
        data.get("father_mobile"),              # 38  Family Contact
        "Father",                               # 39  Family Relation
        "",                                     # 40  Family Age
        data.get("marital_status"),             # 41  Family Marital
        data.get("father_occupation"),          # 42  Family Occupation
        data.get("tenth_school"),               # 43  Exam Name
        data.get("tenth_board"),                # 44  Board
        data.get("tenth_year"),                 # 45  Year
        "",                                     # 46  Subject
        data.get("tenth_percent"),              # 47  Percentage
        # Previous Employment
        data.get("prev_emp_company"),           # 48  Prev Emp Company
        data.get("prev_emp_from"),              # 49  Prev Emp From
        data.get("prev_emp_to"),                # 50  Prev Emp To
        data.get("prev_emp_categories"),        # 51  Prev Emp Categories
        # Foreign Employment
        data.get("foreign_name"),               # 52  Foreign Name
        data.get("foreign_relation"),           # 53  Foreign Relation
        data.get("foreign_job"),                # 54  Foreign Job
        data.get("foreign_country"),            # 55  Foreign Country
        data.get("foreign_duration"),           # 56  Foreign Duration
        "",                                     # 57  blank
        app_id,                                 # 58  APPLICATION ID
        sub_date,                               # 59  Submission Date
        sub_time,                               # 60  Submission Time
        photo_name,                             # 61  Photo File
        pdf_name,                               # 62  Rec Letter File
        gen_pdf_name,                           # 63  Generated PDF
    ]
    wr(ws_app, app_row, fill=fill)

    # ── Sheet: Admin ─────────────────────────────────────
    # Col: 1=SNo, 2=Name, 3=Qual, 4=Branch, 5=DOB, 6=College,
    #      7=Semester, 8=CGPA, 9=From, 10=To, 11=ApplicationID
    ws_adm = wb["Admin"]
    adm_sno = max(0, ws_adm.max_row - 1) + 1
    adm_fill = alt1 if adm_sno % 2 == 1 else alt2
    wr(ws_adm, [
        adm_sno,
        data.get("name"),
        data.get("qualification"),
        data.get("btech_branch"),
        data.get("dob"),
        data.get("college_name"),
        data.get("btech_year"),
        data.get("btech_cgpa"),
        data.get("duration_from"),
        data.get("duration_to"),
        app_id,
    ], fill=adm_fill)

    # ── Sheet: Family Details ────────────────────────────
    # Use _extra_fam list (properly collected in /submit route)
    # Cols: 1=AppID, 2=Name, 3-5=Father, 6-8=Mother,
    #       9-12=Sib1, 13-16=Sib2, 17-20=Sib3, 21-24=Sib4,
    #       25=NumSiblings, 26=SiblingDetails
    ws_fam = wb["Family Details"]
    extra_fam = data.get("_extra_fam", [])

    def sib(n):
        if n <= len(extra_fam):
            s = extra_fam[n-1]
            return [s.get("name",""), s.get("type",""), s.get("occupation",""), s.get("mobile","")]
        return ["","","",""]

    fam_sno = max(0, ws_fam.max_row - 1) + 1
    wr(ws_fam, [
        app_id,                          # 1  App ID
        data.get("name"),                # 2  Applicant Name
        data.get("father_name"),         # 3  Father Name
        data.get("father_occupation"),   # 4  Father Occupation
        data.get("father_mobile"),       # 5  Father Mobile
        data.get("mother_name"),         # 6  Mother Name
        data.get("mother_occupation"),   # 7  Mother Occupation
        data.get("mother_mobile"),       # 8  Mother Mobile
        *sib(1),                         # 9-12  Sibling 1
        *sib(2),                         # 13-16 Sibling 2
        *sib(3),                         # 17-20 Sibling 3
        *sib(4),                         # 21-24 Sibling 4
        data.get("num_siblings"),        # 25 Number of Siblings
        data.get("sibling_info"),        # 26 Sibling Details
    ], fill=alt1 if fam_sno % 2 == 1 else alt2)

    # ── Sheet: Education Details ─────────────────────────
    # Use _extra_edu list (properly collected in /submit route)
    # Cols: 1=AppID, 2=Name, 3=Branch,
    #       4-7=10th, 8-11=Inter, 12-15=Degree, 16-19=Masters, 20-23=PhD
    ws_edu = wb["Education Details"]
    extra_edu = data.get("_extra_edu", [])

    def ecol(inst, board, year, pct):
        return [inst or "", board or "", year or "", pct or ""]

    tenth   = ecol(data.get("tenth_school"),     data.get("tenth_board"),  data.get("tenth_year"),  data.get("tenth_percent"))
    inter   = ecol(data.get("inter_institution"), data.get("inter_board"),  data.get("inter_year"),  data.get("inter_percent"))
    degree  = ecol(data.get("btech_college"),     data.get("btech_branch"), data.get("btech_year"),  data.get("btech_cgpa"))
    masters = ecol(extra_edu[0].get("inst","") if extra_edu else "", extra_edu[0].get("board","") if extra_edu else "", extra_edu[0].get("year","") if extra_edu else "", extra_edu[0].get("percent","") if extra_edu else "")
    phd     = ecol(extra_edu[1].get("inst","") if len(extra_edu)>1 else "", extra_edu[1].get("board","") if len(extra_edu)>1 else "", extra_edu[1].get("year","") if len(extra_edu)>1 else "", extra_edu[1].get("percent","") if len(extra_edu)>1 else "")

    edu_sno = max(0, ws_edu.max_row - 2) + 1
    wr(ws_edu, [
        app_id,                  # 1  App ID
        data.get("name"),        # 2  Name
        data.get("btech_branch"),# 3  Branch
        *tenth,                  # 4-7
        *inter,                  # 8-11
        *degree,                 # 12-15
        *masters,                # 16-19
        *phd,                    # 20-23
    ], fill=alt1 if edu_sno % 2 == 1 else alt2)

    wb.save(EXCEL_PATH)

def email_exists(email):
    if not os.path.exists(EXCEL_PATH):
        return False
    ensure_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    if "Application" not in wb.sheetnames:
        return False
    ws = wb["Application"]
    # Email is in column 22
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[21] and str(row[21]).strip().lower() == email.strip().lower():
            return True
    return False

# ── PDF Generation ──────────────────────────────────────────
BLUE     = colors.HexColor("#1565C0")
DARKBLUE = colors.HexColor("#0d3a6b")
LBLUE    = colors.HexColor("#e3f2fd")
GREY     = colors.HexColor("#f5f5f5")
GREEN    = colors.HexColor("#e8f5e9")
GREENB   = colors.HexColor("#43a047")

def generate_application_pdf(data, app_id, sub_date, sub_time, photo_path, rec_pdf_path):
    """Generate application PDF — spacious layout. Undertaking is a separate downloadable PDF."""
    filename = f"ADA_{app_id}.pdf"
    out_path = os.path.join(UPLOAD_GEN_DIR, filename)

    styles = getSampleStyleSheet()

    # ── Styles ──────────────────────────────────────────
    def S(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    title_s  = S("T",  fontSize=14, fontName="Helvetica-Bold", alignment=TA_CENTER, textColor=DARKBLUE, spaceAfter=4, spaceBefore=4)
    sub_s    = S("S",  fontSize=10, fontName="Helvetica-Bold", alignment=TA_CENTER, textColor=DARKBLUE, spaceAfter=2)
    apptag_s = S("AT", fontSize=11, fontName="Helvetica-Bold", alignment=TA_CENTER, textColor=BLUE, spaceAfter=6)
    sec_s    = S("Sec",fontSize=10, fontName="Helvetica-Bold", alignment=TA_LEFT,   textColor=colors.white)
    body_s   = S("B",  fontSize=10, leading=16, spaceAfter=4)
    lbl_s    = S("L",  fontSize=9,  fontName="Helvetica-Bold", leading=15, textColor=colors.HexColor("#37474f"))
    val_s    = S("V",  fontSize=10, leading=15)
    foot_s   = S("F",  fontSize=8,  alignment=TA_CENTER, textColor=colors.grey)
    note_s   = S("N",  fontSize=9,  leading=14, textColor=colors.HexColor("#555"), alignment=TA_CENTER)

    story = []

    def sp(h=0.4):
        story.append(Spacer(1, h*cm))

    def hr(thick=1.2, c=BLUE):
        story.append(HRFlowable(width="100%", thickness=thick, color=c, spaceAfter=6, spaceBefore=6))

    def section_hdr(num, text):
        tbl = Table([[Paragraph(f"  {num}. {text}", sec_s)]], colWidths=[17.4*cm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), BLUE),
            ("TOPPADDING",(0,0),(-1,-1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1), 8),
            ("LEFTPADDING",(0,0),(-1,-1), 6),
            ("ROUNDEDCORNERS", [4]),
        ]))
        story.append(tbl)
        sp(0.2)

    def info_table(rows, widths=None):
        if not widths:
            widths = [5.8*cm, 11.6*cm]
        tbl_data = []
        for lbl, val in rows:
            tbl_data.append([
                Paragraph(lbl, lbl_s),
                Paragraph(str(val) if val else "—", val_s)
            ])
        t = Table(tbl_data, colWidths=widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0),(0,-1), LBLUE),
            ("GRID",       (0,0),(-1,-1), 0.5, colors.HexColor("#cfd8dc")),
            ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
            ("TOPPADDING", (0,0),(-1,-1), 7),
            ("BOTTOMPADDING",(0,0),(-1,-1), 7),
            ("LEFTPADDING",(0,0),(-1,-1), 8),
            ("ROWBACKGROUNDS",(0,0),(-1,-1), [colors.white, colors.HexColor("#f5f9ff")]),
        ]))
        story.append(t)
        sp(0.3)

    # ── PAGE HEADER with Logo ────────────────────────────
    logo_path = os.path.join(BASE_DIR, "static", "img", "ada_logo.png")
    header_content = []
    if os.path.exists(logo_path):
        try:
            logo_img = RLImage(logo_path, width=2*cm, height=2*cm)
            hdr_tbl = Table([[
                logo_img,
                [Paragraph("AERONAUTICAL DEVELOPMENT AGENCY (ADA)", title_s),
                 Paragraph("BANGALORE — 560 017", sub_s)]
            ]], colWidths=[2.5*cm, 14.9*cm])
            hdr_tbl.setStyle(TableStyle([
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                ("ALIGN",(0,0),(0,0),"CENTER"),
                ("LEFTPADDING",(1,0),(1,0),10),
                ("TOPPADDING",(0,0),(-1,-1),0),
                ("BOTTOMPADDING",(0,0),(-1,-1),0),
            ]))
            story.append(hdr_tbl)
        except Exception:
            story.append(Paragraph("AERONAUTICAL DEVELOPMENT AGENCY (ADA)", title_s))
            story.append(Paragraph("BANGALORE — 560 017", sub_s))
    else:
        story.append(Paragraph("AERONAUTICAL DEVELOPMENT AGENCY (ADA)", title_s))
        story.append(Paragraph("BANGALORE — 560 017", sub_s))
    hr(2)
    story.append(Paragraph("INTERNSHIP / PROJECT WORK — APPLICATION FORM", apptag_s))
    sp(0.2)

    # App ID banner
    banner = Table([[
        Paragraph(f"<b>Application ID:</b>  {app_id}", body_s),
        Paragraph(f"<b>Date:</b>  {sub_date}", body_s),
        Paragraph(f"<b>Time:</b>  {sub_time}", body_s),
    ]], colWidths=[7*cm, 5*cm, 5.4*cm])
    banner.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), GREEN),
        ("BOX",(0,0),(-1,-1), 1.2, GREENB),
        ("TOPPADDING",(0,0),(-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ("LEFTPADDING",(0,0),(-1,-1), 10),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(banner)
    sp(0.5)

    # ── 1. PERSONAL DETAILS + PHOTO ─────────────────────
    section_hdr(1, "PERSONAL DETAILS")

    pers_rows = [
        ("Full Name",      data.get("name")),
        ("Gender",         data.get("gender")),
        ("Date of Birth",  data.get("dob")),
        ("Aadhaar Number", data.get("aadhaar")),
        ("Mobile Number",  data.get("mobile")),
        ("WhatsApp No.",   data.get("whatsapp")),
        ("Email ID",       data.get("email")),
        ("Qualification",  data.get("qualification")),
    ]

    if photo_path and os.path.exists(photo_path):
        try:
            photo_el = RLImage(photo_path, width=3.3*cm, height=4*cm)
        except:
            photo_el = Paragraph("[ Photo ]", body_s)
    else:
        photo_el = Paragraph("[ No Photo ]", body_s)

    pers_tbl = []
    for lbl, val in pers_rows:
        pers_tbl.append([Paragraph(lbl, lbl_s), Paragraph(str(val) if val else "—", val_s)])

    inner = Table(pers_tbl, colWidths=[4.8*cm, 8.5*cm])
    inner.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,-1), LBLUE),
        ("GRID",(0,0),(-1,-1), 0.5, colors.HexColor("#cfd8dc")),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1), 7),
        ("BOTTOMPADDING",(0,0),(-1,-1), 7),
        ("LEFTPADDING",(0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS",(0,0),(-1,-1), [colors.white, colors.HexColor("#f5f9ff")]),
    ]))

    photo_wrapper = Table([[photo_el]], colWidths=[3.8*cm])
    photo_wrapper.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),1,colors.HexColor("#b0bec5")),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),6),
        ("BOTTOMPADDING",(0,0),(-1,-1),6),
    ]))

    outer = Table([[inner, photo_wrapper]], colWidths=[13.8*cm, 3.8*cm])
    outer.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",(1,0),(1,0), 10),
    ]))
    story.append(outer)
    sp(0.2)

    info_table([
        ("Branch",         data.get("btech_branch")),
        ("Year / Semester",data.get("btech_year")),
        ("CGPA",           data.get("btech_cgpa")),
        ("Marital Status", data.get("marital_status")),
        ("Health Issues", data.get("health_issues")),
        ("Designation",    data.get("designation")),
    ])

    # ── 2. ADDRESS DETAILS ───────────────────────────────
    section_hdr(2, "ADDRESS DETAILS")
    info_table([
        ("Permanent Address", data.get("permanent_address")),
        ("Present Address",   data.get("present_address")),
        ("Previous Address",  data.get("previous_address")),
    ])

    # ── 3. CONTACT & IDENTITY ────────────────────────────
    section_hdr(3, "CONTACT & IDENTITY DETAILS")
    info_table([
        ("Mobile",     data.get("mobile")),
        ("WhatsApp",   data.get("whatsapp")),
        ("Landline",   data.get("landline")),
        ("IMEI No.",   data.get("imei")),
        ("PAN No.",    data.get("pan")),
        ("Facebook",   data.get("facebook")),
        ("LinkedIn",   data.get("linkedin")),
        ("Instagram",  data.get("instagram")),
        ("Other ID",   data.get("other_id")),
    ])

    # ── 4. BANK DETAILS ──────────────────────────────────
    section_hdr(4, "BANK DETAILS")
    info_table([
        ("Bank Name",    data.get("bank_name")),
        ("Account No.",  data.get("bank_account")),
        ("IFSC Code",    data.get("ifsc")),
        ("EPF No.",      data.get("epf")),
        ("ESIC No.",     data.get("esic")),
    ])

    # ── 5. EDUCATION DETAILS (table like the screenshot) ─
    section_hdr(5, "EDUCATION DETAILS")

    # Collect all edu entries — use _extra_edu list populated in /submit
    edu_levels = [
        ("10th Standard",
         data.get("tenth_school"), data.get("tenth_board"),
         data.get("tenth_year"),   data.get("tenth_percent")),
        ("Intermediate/Diploma",
         data.get("inter_institution"), data.get("inter_board"),
         data.get("inter_year"),        data.get("inter_percent")),
        ("B.Tech/Degree",
         (data.get("btech_college") or "") + ("\n" + data.get("btech_branch") if data.get("btech_branch") else ""),
         "—", data.get("btech_year"), data.get("btech_cgpa")),
    ]
    # Append dynamically added rows from _extra_edu list
    for e in data.get("_extra_edu", []):
        edu_levels.append((
            e.get("level") or "Additional",
            e.get("inst") or "—",
            e.get("board") or "—",
            e.get("year") or "—",
            e.get("percent") or "—",
        ))

    edu_hdr = [
        Paragraph("<b>Level</b>", lbl_s),
        Paragraph("<b>School/Institution</b>", lbl_s),
        Paragraph("<b>Board/University</b>", lbl_s),
        Paragraph("<b>Year</b>", lbl_s),
        Paragraph("<b>Percentage/CGPA</b>", lbl_s),
    ]
    edu_data = [edu_hdr]
    for lev, inst, board, year, pct in edu_levels:
        edu_data.append([
            Paragraph(str(lev) if lev else "—", val_s),
            Paragraph(str(inst) if inst else "—", val_s),
            Paragraph(str(board) if board else "—", val_s),
            Paragraph(str(year) if year else "—", val_s),
            Paragraph(str(pct) if pct else "—", val_s),
        ])

    edu_t = Table(edu_data, colWidths=[3.5*cm, 5.2*cm, 3.5*cm, 2*cm, 3.2*cm])
    edu_t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0), BLUE),
        ("TEXTCOLOR",(0,0),(-1,0), colors.white),
        ("GRID",(0,0),(-1,-1), 0.5, colors.HexColor("#cfd8dc")),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ("LEFTPADDING",(0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#e3f2fd")]),
    ]))
    story.append(edu_t)
    sp(0.3)

    # ── 6. FAMILY DETAILS (table like screenshot) ────────
    section_hdr(6, "FAMILY DETAILS")

    # Collect all family members — use _extra_fam list populated in /submit
    fam_members = [
        ("Father", data.get("father_name"), data.get("father_occupation"), data.get("father_mobile")),
        ("Mother", data.get("mother_name"), data.get("mother_occupation"), data.get("mother_mobile")),
    ]
    for m in data.get("_extra_fam", []):
        fam_members.append((
            m.get("type") or "Member",
            m.get("name") or "—",
            m.get("occupation") or "—",
            m.get("mobile") or "—",
        ))

    fam_hdr = [
        Paragraph("", lbl_s),
        Paragraph("<b>Name</b>", lbl_s),
        Paragraph("<b>Occupation</b>", lbl_s),
        Paragraph("<b>Mobile Number</b>", lbl_s),
    ]
    fam_data = [fam_hdr]
    for rel, name, occ, mob in fam_members:
        fam_data.append([
            Paragraph(f"<b>{rel}</b>", lbl_s),
            Paragraph(str(name) if name else "—", val_s),
            Paragraph(str(occ) if occ else "—", val_s),
            Paragraph(str(mob) if mob else "—", val_s),
        ])

    if data.get("num_siblings") or data.get("sibling_info"):
        fam_data.append([
            Paragraph("<b>Siblings</b>", lbl_s),
            Paragraph(f"Count: {data.get('num_siblings') or '—'}", val_s),
            Paragraph(data.get("sibling_info") or "—", val_s),
            Paragraph("", val_s),
        ])

    fam_t = Table(fam_data, colWidths=[2.5*cm, 5.5*cm, 5*cm, 4.4*cm])
    fam_t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0), BLUE),
        ("TEXTCOLOR",(0,0),(-1,0), colors.white),
        ("GRID",(0,0),(-1,-1), 0.5, colors.HexColor("#cfd8dc")),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ("LEFTPADDING",(0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#e3f2fd")]),
    ]))
    story.append(fam_t)
    sp(0.3)

    # ── 7. PROJECT DETAILS ───────────────────────────────
    section_hdr(7, "PROJECT DETAILS")
    info_table([
        ("Project Title",    data.get("project_title")),
        ("Guide/Directorate",data.get("guide")),
        ("Area of Work",     data.get("area_of_work")),
        ("Duration From",    data.get("duration_from")),
        ("Duration To",      data.get("duration_to")),
        ("University Reg No",data.get("university_reg")),
    ])

    # ── 8. FACULTY COORDINATOR ───────────────────────────
    section_hdr(8, "FACULTY COORDINATOR DETAILS")
    info_table([
        ("Faculty Name",    data.get("faculty_name")),
        ("Designation",     data.get("faculty_designation")),
        ("Department",      data.get("faculty_department")),
        ("Affiliation ID",  data.get("affiliation_id")),
        ("Email ID",        data.get("faculty_email")),
        ("Contact No.",     data.get("faculty_contact")),
        ("Fax No.",         data.get("faculty_fax")),
    ])

    # ── 9. COLLEGE DETAILS ───────────────────────────────
    section_hdr(9, "COLLEGE DETAILS")
    info_table([
        ("College Name",               data.get("college_name")),
        ("Principal Name",             data.get("principal_name")),
        ("University Affiliation",     data.get("university_affiliation_name")),
        ("Affiliation No.",            data.get("university_affiliation_no")),
        ("AICTE Permanent Code",       data.get("aicte_code")),
        ("DTE Code",                   data.get("dte_code")),
        ("College Email",              data.get("college_email")),
        ("Contact Number",             data.get("college_contact")),
        ("Fax Number",                 data.get("college_fax")),
    ])

    # ── 10. PREVIOUS EMPLOYMENT ─────────────────────────
    section_hdr(10, "PREVIOUS EMPLOYMENT DETAILS")
    info_table([
        ("Company Name",       data.get("prev_emp_company")),
        ("From",               data.get("prev_emp_from")),
        ("To",                 data.get("prev_emp_to")),
        ("Categories of Work", data.get("prev_emp_categories")),
    ])

    # ── 11. FOREIGN EMPLOYMENT ───────────────────────────
    section_hdr(11, "FOREIGN EMPLOYMENT DETAILS (Self / Family / Relatives)")
    info_table([
        ("Name",     data.get("foreign_name")),
        ("Relation", data.get("foreign_relation")),
        ("Job",      data.get("foreign_job")),
        ("Country",  data.get("foreign_country")),
        ("Duration", data.get("foreign_duration")),
    ])

    sp(0.4)
    hr(1.5)
    story.append(Paragraph(
        f"Application ID: {app_id}  ·  Submitted on {sub_date} at {sub_time}  ·  Aeronautical Development Agency, Bangalore 560017",
        foot_s))
    sp(0.15)
    story.append(Paragraph(
        "Note: Download and print the ADA Undertaking Form separately, get it signed by your HOD/Principal, and submit it on reporting day.",
        note_s))

    doc = SimpleDocTemplate(out_path, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm)
    doc.build(story)

    # ── Merge: app PDF + uploaded rec letter PDF + photo page ──
    try:
        merged_path = out_path.replace(".pdf", "_FULL.pdf")
        with pikepdf.Pdf.new() as merged:
            # 1. Application details pages
            with pikepdf.open(out_path) as app_pdf:
                merged.pages.extend(app_pdf.pages)

            # 2. Uploaded recommendation letter PDF
            if rec_pdf_path and os.path.exists(rec_pdf_path):
                try:
                    with pikepdf.open(rec_pdf_path) as rec_pdf:
                        merged.pages.extend(rec_pdf.pages)
                except Exception:
                    pass  # skip if rec letter can't be opened

            # 3. Photo page — embed photo in a mini PDF page
            if photo_path and os.path.exists(photo_path):
                try:
                    photo_page_path = out_path.replace(".pdf", "_photopage.pdf")
                    _make_photo_page(photo_page_path, photo_path, data.get("name",""), app_id)
                    with pikepdf.open(photo_page_path) as ph_pdf:
                        merged.pages.extend(ph_pdf.pages)
                    os.remove(photo_page_path)
                except Exception:
                    pass

            merged.save(merged_path, compress_streams=True,
                        object_stream_mode=pikepdf.ObjectStreamMode.generate)

        # Replace the base PDF with the merged version
        os.replace(merged_path, out_path)
    except Exception:
        pass  # if merge fails, the base application PDF is still valid

    return filename

def _make_photo_page(out_path, photo_path, name, app_id):
    """Create a single A4 PDF page containing the applicant photo + label."""
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("PT", parent=styles["Normal"], fontSize=12,
                              fontName="Helvetica-Bold", alignment=TA_CENTER,
                              textColor=colors.HexColor("#1565C0"), spaceAfter=8)
    sub_s   = ParagraphStyle("PS", parent=styles["Normal"], fontSize=10,
                              alignment=TA_CENTER, textColor=colors.HexColor("#37474f"), spaceAfter=4)

    story = []
    story.append(Spacer(1, 2*cm))
    story.append(Paragraph("AERONAUTICAL DEVELOPMENT AGENCY (ADA)", title_s))
    story.append(Paragraph("APPLICANT PHOTOGRAPH", title_s))
    story.append(Spacer(1, 0.5*cm))

    try:
        img = RLImage(photo_path, width=6*cm, height=7.5*cm)
        photo_tbl = Table([[img]], colWidths=[7*cm])
        photo_tbl.setStyle(TableStyle([
            ("BOX",(0,0),(-1,-1), 1.5, colors.HexColor("#1565C0")),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("TOPPADDING",(0,0),(-1,-1), 10),
            ("BOTTOMPADDING",(0,0),(-1,-1), 10),
            ("LEFTPADDING",(0,0),(-1,-1), 10),
            ("RIGHTPADDING",(0,0),(-1,-1), 10),
            ("BACKGROUND",(0,0),(-1,-1), colors.HexColor("#f5f9ff")),
        ]))
        story.append(photo_tbl)
    except Exception:
        story.append(Paragraph("[Photo not available]", sub_s))

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"<b>Name:</b> {name}", sub_s))
    story.append(Paragraph(f"<b>Application ID:</b> {app_id}", sub_s))

    doc = SimpleDocTemplate(out_path, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm)
    doc.build(story)


@app.route("/")
def index():
    return render_template("index.html")

@app.route("/apply")
def apply():
    return render_template("apply.html")

@app.route("/submit", methods=["POST"])
def submit():
    f     = request.form
    email = f.get("email", "").strip()

    if email_exists(email):
        return jsonify({"success": False,
                        "error": "This email ID is already registered. Each applicant may submit only once."})

    rec_file   = request.files.get("rec_letter")
    photo_file = request.files.get("photo")


    if not rec_file or rec_file.filename == "":
        return jsonify({"success": False, "error": "Please upload the Recommendation Letter PDF."})
    if not photo_file or photo_file.filename == "":
        return jsonify({"success": False, "error": "Please upload your passport-size photo."})
    if not allowed_file(rec_file.filename, ALLOWED_PDF):
        return jsonify({"success": False, "error": "Recommendation letter must be a .pdf file."})
    if not allowed_file(photo_file.filename, ALLOWED_IMG):
        return jsonify({"success": False, "error": "Photo must be JPG, JPEG or PNG."})

    # Compress and save photo — hard reject if too large
    photo_file.stream.seek(0, 2)
    photo_size_kb = photo_file.stream.tell() / 1024
    photo_file.stream.seek(0)
    if photo_size_kb > 100:
        return jsonify({"success": False, "error": f"Photo must be under 100 KB. Your file is {photo_size_kb:.0f} KB."})
    photo_name = unique_filename("jpg")
    photo_path = os.path.join(UPLOAD_PHO_DIR, photo_name)
    photo_file.save(photo_path)

    # Compress and save recommendation letter — hard reject if too large
    rec_file.stream.seek(0, 2)
    pdf_size_kb = rec_file.stream.tell() / 1024
    rec_file.stream.seek(0)
    if pdf_size_kb > 300:
        return jsonify({"success": False, "error": f"Recommendation letter must be under 300 KB. Your file is {pdf_size_kb:.0f} KB."})
    pdf_name = unique_filename("pdf")
    pdf_path = os.path.join(UPLOAD_PDF_DIR, pdf_name)
    rec_file.save(pdf_path)

    
    fields = [
        "name","gender","dob","aadhaar","mobile","email","whatsapp","landline","imei",
        "permanent_address","present_address","previous_address",
        "qualification","designation","marital_status","health_issues",
        "pan","epf","esic","bank_name","bank_account","ifsc",
        "facebook","linkedin","instagram","other_id",
        "project_title","guide","area_of_work","duration_from","duration_to",
        "university_reg",# Previous Employment
        "prev_emp_company","prev_emp_from","prev_emp_to","prev_emp_categories",
        # Foreign Employment
        "foreign_name","foreign_relation","foreign_job","foreign_country","foreign_duration",
        # Education — standard fields
        "tenth_school","tenth_board","tenth_year","tenth_percent",
        "inter_institution","inter_board","inter_year","inter_percent",
        "btech_college","btech_branch","btech_cgpa","btech_year",
        # Family — standard fields
        "father_name","father_occupation","father_mobile",
        "mother_name","mother_occupation","mother_mobile",
        "num_siblings","sibling_info",
        # Faculty
        "faculty_name","faculty_designation","faculty_department","affiliation_id",
        "faculty_email","faculty_contact","faculty_fax",
        # College
        "college_name","principal_name","university_affiliation_name","university_affiliation_no",
        "aicte_code","dte_code","college_email","college_contact","college_fax",
    ]
    data = {}
    for k in request.form:
        data[k] = request.form.get(k, "").strip()
      
    data["email"] = email

    # Collect ALL dynamic education add-rows (edu_level_1, edu_inst_1 … up to 20)
    data["_extra_edu"] = []
    for idx in range(1, 21):
        elevel = f.get(f"edu_level_{idx}", "").strip()
        einst  = f.get(f"edu_inst_{idx}", "").strip()
        eboard = f.get(f"edu_board_{idx}", "").strip()
        eyear  = f.get(f"edu_year_{idx}", "").strip()
        epct   = f.get(f"edu_percent_{idx}", "").strip()
        if not elevel and not einst:
            break
        data["_extra_edu"].append({
            "level": elevel, "inst": einst,
            "board": eboard, "year": eyear, "percent": epct
        })
        # Also store flat for backwards compat
        data[f"edu_level_{idx}"] = elevel or None
        data[f"edu_inst_{idx}"]  = einst or None
        data[f"edu_board_{idx}"] = eboard or None
        data[f"edu_year_{idx}"]  = eyear or None
        data[f"edu_percent_{idx}"] = epct or None

    # Collect ALL dynamic family add-rows (fam_type_1, fam_name_1 … up to 20)
    data["_extra_fam"] = []
    for idx in range(1, 21):
        mtype = f.get(f"fam_type_{idx}", "").strip()
        mname = f.get(f"fam_name_{idx}", "").strip()
        mocc  = f.get(f"fam_occupation_{idx}", "").strip()
        mmob  = f.get(f"fam_mobile_{idx}", "").strip()
        if not mtype and not mname:
            break
        data["_extra_fam"].append({
            "type": mtype, "name": mname, "occupation": mocc, "mobile": mmob
        })
        data[f"fam_type_{idx}"]       = mtype or None
        data[f"fam_name_{idx}"]       = mname or None
        data[f"fam_occupation_{idx}"] = mocc or None
        data[f"fam_mobile_{idx}"]     = mmob or None

    # Generate App ID + timestamp
    app_id   = generate_app_id()
    from datetime import timezone, timedelta
    IST = timezone(timedelta(hours=5, minutes=30))
    now      = datetime.now(IST)
    sub_date = now.strftime("%d-%m-%Y")
    hour     = now.hour
    am_pm    = "AM" if hour < 12 else "PM"
    hour12   = hour % 12 or 12
    sub_time = f"{hour12:02d}:{now.minute:02d}:{now.second:02d} {am_pm}"

    # Generate consolidated PDF
    try:
        gen_pdf_name = generate_application_pdf(data, app_id, sub_date, sub_time, photo_path, pdf_path)
    except Exception as e:
        gen_pdf_name = None

    # Save to Excel
    try:
        save_to_excel(data, app_id, sub_date, sub_time, photo_name, pdf_name, gen_pdf_name or "")
    except Exception as e:
        return jsonify({"success": False, "error": f"Excel save error: {str(e)}"})

    return jsonify({
        "success": True,
        "app_id": app_id,
        "sub_date": sub_date,
        "sub_time": sub_time,
        "pdf_available": gen_pdf_name is not None
    })

@app.route("/success")
def success():
    return render_template("success.html")

@app.route("/download_pdf/<app_id>")
def download_pdf(app_id):
    """Allow applicant to download their generated PDF by app_id."""
    # Sanitize
    safe_id = "".join(c for c in app_id if c.isalnum())
    filename = f"ADA_{safe_id}.pdf"
    path = os.path.join(UPLOAD_GEN_DIR, filename)
    if not os.path.exists(path):
        return "PDF not found. Please contact admin.", 404
    return send_from_directory(UPLOAD_GEN_DIR, filename, as_attachment=True,
                                download_name=filename)

@app.route("/redownload", methods=["GET", "POST"])
def redownload():
    """Page for applicants to re-download their PDF by App ID."""
    if request.method == "POST":
        app_id = request.form.get("app_id", "").strip()
        safe_id = "".join(c for c in app_id if c.isalnum())
        filename = f"ADA_{safe_id}.pdf"
        path = os.path.join(UPLOAD_GEN_DIR, filename)
        if os.path.exists(path):
            return send_from_directory(UPLOAD_GEN_DIR, filename, as_attachment=True,
                                        download_name=filename)
        else:
            return render_template("redownload.html", error="Application ID not found. Please check and try again.")
    return render_template("redownload.html")

@app.route("/download_undertaking")
def download_undertaking():
    """Serve the original ADA Undertaking Form PDF for applicants to download, print and sign."""
    # Prefer the real PDF uploaded by admin
    pdf_path = os.path.join(SAMPLE_DIR, "ADA_Undertaking_Form.pdf")
    if os.path.exists(pdf_path):
        return send_from_directory(SAMPLE_DIR, "ADA_Undertaking_Form.pdf",
            as_attachment=True, download_name="ADA_Undertaking_Form.pdf")
    # Fallback to docx if pdf not present
    docx_path = os.path.join(SAMPLE_DIR, "ADA_Undertaking_Form.docx")
    if os.path.exists(docx_path):
        return send_from_directory(SAMPLE_DIR, "ADA_Undertaking_Form.docx",
            as_attachment=True, download_name="ADA_Undertaking_Form.docx")
    return "Undertaking form not available. Please contact admin.", 404

# ── Admin Auth ────────────────────────────────────────────────
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        em  = request.form.get("email", "").strip()
        pwd = request.form.get("password", "").strip()
        if em == ADMIN_EMAIL and hashlib.sha256(pwd.encode()).hexdigest() == ADMIN_PASS_HASH:
            session["admin_logged_in"] = True
            return redirect(url_for("admin"))
        flash("Invalid email or password.", "danger")
    return render_template("admin_login.html")

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("admin_login"))

@app.route("/admin")
def admin():
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login"))

    search  = request.args.get("search", "").strip().lower()
    branch  = request.args.get("branch", "")
    applicants = []
    branches   = []

    ensure_excel()
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Application" in wb.sheetnames:
            ws = wb["Application"]
            seen_branches = set()
            # Row 1 is the header; data starts at row 2
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[57]:   # APPLICATION ID (col 58, idx 57)
                    continue
                a = {
                    "App ID":        row[57],   # col 58
                    "Name":          row[1],
                    "Gender":        row[10],
                    "DOB":           row[4],
                    "Qualification": row[2],
                    "Branch":        row[3],
                    "College":       row[5],
                    "Semester":      row[6],
                    "CGPA":          row[7],
                    "Email":         row[20],
                    "Mobile":        row[15],
                    "Aadhaar":       row[9],
                    "Health Issues": row[12],
                    "Duration From": row[33],
                    "Duration To":   row[34],
                    "Sub Date":      row[58],   # col 59
                    "Sub Time":      row[59],   # col 60
                    "Photo File":    row[60],   # col 61
                    "PDF File":      row[61],   # col 62
                    "Gen PDF":       row[62],   # col 63
                }
                b = a["Branch"] or ""
                if b and b not in seen_branches:
                    branches.append(b); seen_branches.add(b)
                if search and search not in str(a.get("Name","")).lower() \
                           and search not in str(a.get("App ID","")).lower() \
                           and search not in str(a.get("Email","")).lower():
                    continue
                if branch and a.get("Branch","") != branch:
                    continue
                applicants.append(a)

    return render_template("admin.html",
        applicants=applicants, total=len(applicants),
        branches=branches, search=search, selected_branch=branch)

@app.route("/admin/download_excel")
def download_excel():
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login"))
    return send_from_directory(BASE_DIR, "APPLICATION.xlsx", as_attachment=True,
                                download_name="ADA_Applications.xlsx")

@app.route("/admin/download_gen_pdf/<app_id>")
def admin_download_gen_pdf(app_id):
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login"))
    safe_id  = "".join(c for c in app_id if c.isalnum())
    filename = f"ADA_{safe_id}.pdf"
    path = os.path.join(UPLOAD_GEN_DIR, filename)
    if not os.path.exists(path):
        return "PDF not found.", 404
    return send_from_directory(UPLOAD_GEN_DIR, filename, as_attachment=True)

@app.route("/uploads/pdfs/<filename>")
def serve_pdf(filename):
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login"))
    return send_from_directory(UPLOAD_PDF_DIR, filename)

@app.route("/uploads/photos/<filename>")
def serve_photo(filename):
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login"))
    return send_from_directory(UPLOAD_PHO_DIR, filename)

# ── Run ───────────────────────────────────────────────────────
if __name__ == "__main__":
    ensure_excel()
    port  = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV") != "production"
    app.run(host="0.0.0.0", port=port, debug=debug)
